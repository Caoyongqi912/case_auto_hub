"""
M1/M2 模板上传 适用端 (case_platform) 必填 + 枚举校验 回归测试.

锁定行为:
- M1 路径 (utils.aioFileReader._read): 空值只报 1 次 "不能为空", 非法 label 报
  "不在可选范围内", 枚举空时不卡 enum (跟 M1 用例等级的语义对齐).
- M2 commit 路径 (M2ImportService._validate_field_constraints): 解析阶段故意
  enum_config=空, commit 入口补卡; 校验通过把 label 替换为 value, 跟 M1 _resolve_enum
  行为对齐.
"""
from io import BytesIO
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest

from app.exception import CommonError
from app.service.m2ImportService import M2ImportService
from utils.aioFileReader import AsyncFilesReader, CaseEnumConfig


# 测试用的固定枚举 (避免依赖真实 case_config 表)
LEVEL_MAP = {"P0": "P0", "P1": "P1", "P2": "P2"}
TYPE_MAP = {"功能": "GN", "冒烟": "MY"}
PLATFORM_MAP = {"PC端": "PC", "H5": "H5", "APP": "APP", "WJAPP端": "WJAPP"}

VALID_ENUM = CaseEnumConfig(
    level_map=LEVEL_MAP,
    type_map=TYPE_MAP,
    platform_map=PLATFORM_MAP,
    default_level=None,
    default_type=None,
    default_platform=None,
)

EMPTY_ENUM = CaseEnumConfig()  # 跟 M2 解析阶段传的空 config 等价


def _build_xlsx(rows: list[dict]) -> bytes:
    """
    构造一个最简 M1 模板, 写一行表头 + N 行数据.
    列顺序对齐 FIELD_MAPPING (utils/aioFileReader.py:41).
    """
    headers = [
        "标题*", "前置条件", "步骤描述", "预期结果",
        "用例等级*", "用例类型", "适用端", "备注", "所属分组",
    ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用例"
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("case_name", "测试用例"),
            r.get("case_setup", ""),
            r.get("action", "1. 打开页面"),
            r.get("expected_result", "1. 成功"),
            r.get("case_level", "P0"),
            r.get("case_type", "功能"),
            r.get("case_platform", "PC端"),
            r.get("case_mark", ""),
            r.get("group_path", "默认分组"),
        ])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _read_with_enum(content: bytes, enum_config: CaseEnumConfig):
    """M1 路径: 调 AsyncFilesReader._read (内部用 file-like 走 read_excel_async).
    _read 自身是同步, 但被 async_read_from_bytes 包装; 这里走 _read 直测, 避免
    引入 docx/Excel 异步 IO 路径, 测试聚焦在字段校验逻辑.
    """
    reader = AsyncFilesReader(enum_config=enum_config)
    return reader._read(  # noqa: SLF001
        file=BytesIO(content),
        file_md5="test-md5",
        enum_config=enum_config,
    )


# --------------------------------------------------------------------------- #
# M1 路径: utils/aioFileReader._read 适用端校验
# --------------------------------------------------------------------------- #

@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_empty_platform_reports_single_required_error():
    """适用端为空 -> 1 条 '不能为空' 错误, 不会双报错."""
    content = _build_xlsx([{"case_platform": None}])
    result = _read_with_enum(content, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert len(platform_errs) == 1, (
        f"应只 1 条'不能为空', 实际 {len(platform_errs)} 条: {platform_errs}"
    )
    assert "不能为空" in platform_errs[0]["message"]


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_valid_platform_passes():
    """适用端 = 合法 label (PC端) -> 通过, 转换为 value (PC) 入 mapped_case."""
    content = _build_xlsx([{"case_platform": "PC端"}])
    result = _read_with_enum(content, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert platform_errs == [], f"不应报错, 实际: {platform_errs}"
    assert result.valid_cases, "valid_cases 应有数据"
    # 验证 label -> value 转换 (跟 enum_map 一致)
    assert result.valid_cases[0]["case_platform"] == PLATFORM_MAP["PC端"]


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_invalid_platform_reports_enum_error():
    """适用端 = 非法 label (XBOX) -> 1 条 '不在可选范围内' 错误, 不再追加'不能为空'."""
    content = _build_xlsx([{"case_platform": "XBOX"}])
    result = _read_with_enum(content, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert len(platform_errs) == 1, (
        f"应只 1 条 enum 错误, 实际 {len(platform_errs)} 条: {platform_errs}"
    )
    assert "不在可选范围内" in platform_errs[0]["message"]


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_empty_platform_map_only_validates_required():
    """platform_map 为空 (预览模式 / 配置未加载) -> 空值仍报'不能为空', 非空放行."""
    # 1) 空值
    content = _build_xlsx([{"case_platform": None}])
    result = _read_with_enum(content, EMPTY_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert len(platform_errs) == 1
    assert "不能为空" in platform_errs[0]["message"]

    # 2) 非空值 -> 不查 enum, 放行 (跟 M1 用例等级 else 分支对齐)
    content = _build_xlsx([{"case_platform": "任意端"}])
    result = _read_with_enum(content, EMPTY_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert platform_errs == [], f"enum_map 空时不应查 enum, 实际: {platform_errs}"


# --------------------------------------------------------------------------- #
# M2 commit 路径: M2ImportService._validate_field_constraints
# --------------------------------------------------------------------------- #

@pytest.fixture
def patch_enum_config(monkeypatch):
    """
    mock utils.caseEnumResolver.load_case_enum_config, 避免连真实 DB.
    配合 parametrize 切换 platform_map 状态.
    """
    def _patch(platform_map):
        cfg = CaseEnumConfig(platform_map=platform_map)
        async def _fake_load():
            return cfg
        monkeypatch.setattr(
            "utils.caseEnumResolver.load_case_enum_config",
            _fake_load,
        )
    return _patch


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_empty_platform_raises_required_error(patch_enum_config):
    """M2 commit 校验: 适用端为空 -> CommonError '不能为空'."""
    patch_enum_config(PLATFORM_MAP)
    cases = [{"case_platform": None, "_row": 2}]
    with pytest.raises(CommonError) as exc_info:
        await M2ImportService._validate_field_constraints(cases)
    err = exc_info.value
    assert "不能为空" in str(err.message)
    assert err.data["field"] == "适用端"
    assert err.data["total"] == 1
    assert err.data["errors"][0]["row"] == 2


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_valid_label_converted_to_value(patch_enum_config):
    """M2 commit 校验: 合法 label -> 原地替换为 value, 不报错."""
    patch_enum_config(PLATFORM_MAP)
    cases = [{"case_platform": "PC端", "_row": 2}]
    # 不应抛
    await M2ImportService._validate_field_constraints(cases)
    assert cases[0]["case_platform"] == PLATFORM_MAP["PC端"], (
        f"label 应替换为 value, 实际 {cases[0]['case_platform']!r}"
    )


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_invalid_label_raises_enum_error(patch_enum_config):
    """M2 commit 校验: 非法 label -> CommonError '不在可选范围内'."""
    patch_enum_config(PLATFORM_MAP)
    cases = [{"case_platform": "XBOX", "_row": 3}]
    with pytest.raises(CommonError) as exc_info:
        await M2ImportService._validate_field_constraints(cases)
    err = exc_info.value
    assert "不在可选范围内" in str(err.message)
    assert "XBOX" in str(err.message)
    # 非法 label 不应被改写
    assert cases[0]["case_platform"] == "XBOX"


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_empty_platform_map_only_validates_required(patch_enum_config):
    """M2 commit 校验: platform_map 为空 -> 空值仍报'不能为空', 非空放行不转换."""
    patch_enum_config({})
    # 1) 空值
    with pytest.raises(CommonError) as exc_info:
        await M2ImportService._validate_field_constraints(
            [{"case_platform": None, "_row": 2}]
        )
    assert "不能为空" in str(exc_info.value.message)

    # 2) 非空 -> 不查 enum, 不转换, 放行
    cases = [{"case_platform": "任意端", "_row": 2}]
    await M2ImportService._validate_field_constraints(cases)
    assert cases[0]["case_platform"] == "任意端", (
        "platform_map 为空时不应做 label->value 转换, 保持原值"
    )


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_empty_list_is_noop(patch_enum_config):
    """M2 commit 校验: valid_cases 为空 -> 早返回, 不调 load_case_enum_config."""
    patch_enum_config(PLATFORM_MAP)
    # 不应抛
    await M2ImportService._validate_field_constraints([])


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_multiple_errors_aggregate_with_cap(patch_enum_config):
    """M2 commit 校验: 累计多条错误, 错误信息预览前 10 条 + 提示剩余."""
    patch_enum_config(PLATFORM_MAP)
    cases = [{"case_platform": None, "_row": i} for i in range(2, 15)]  # 13 条
    with pytest.raises(CommonError) as exc_info:
        await M2ImportService._validate_field_constraints(cases)
    err = exc_info.value
    assert err.data["total"] == 13
    assert len(err.data["errors"]) == 13
    assert "还有 3 条" in str(err.message), (
        f"应提示还有 3 条 (13 - 10), 实际: {err.message!r}"
    )



# --------------------------------------------------------------------------- #
# 多值 (case_platform 用逗号分隔, "PC,WJAPP") 回归测试
# --------------------------------------------------------------------------- #

@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_multi_platform_labels_join_with_value_conversion():
    """M1 解析: cell = "PC端, WJAPP端" -> mapped_case["case_platform"] = "PC,WJAPP"."""
    content_bytes = _build_xlsx([{"case_platform": "PC端, WJAPP端"}])
    result = _read_with_enum(content_bytes, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert platform_errs == [], f"多选合法值不应报错, 实际: {platform_errs}"
    assert result.valid_cases, "应有 1 条 valid_case"
    # label 列表按 cell 出现顺序转 value, 保持顺序, 逗号拼接
    assert result.valid_cases[0]["case_platform"] == "PC,WJAPP", (
        f"应转为 'PC,WJAPP', 实际 {result.valid_cases[0]['case_platform']!r}"
    )


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_multi_platform_with_invalid_part_reports_error():
    """M1 解析: cell = "PC端, XBOX端" -> 1 条 enum 错误, 列出非法 label."""
    content_bytes = _build_xlsx([{"case_platform": "PC端, XBOX端"}])
    result = _read_with_enum(content_bytes, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert len(platform_errs) == 1, f"应 1 条错误, 实际: {platform_errs}"
    msg = platform_errs[0]["message"]
    assert "XBOX端" in msg
    assert "不在可选范围内" in msg
    # 不做 partial save: 该行不进 valid_cases
    assert result.valid_cases == []


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_multi_platform_dedup_preserves_first_occurrence():
    """M1 解析: cell = "PC端, WJAPP端, PC端" -> 去重 -> "PC,WJAPP" (保持首次出现顺序)."""
    content_bytes = _build_xlsx([{"case_platform": "PC端, WJAPP端, PC端"}])
    result = _read_with_enum(content_bytes, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert platform_errs == []
    assert result.valid_cases[0]["case_platform"] == "PC,WJAPP"


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_multi_platform_empty_parts_treated_as_empty():
    """M1 解析: cell = "PC端, , WJAPP端" -> 跳过空段, 结果 "PC,WJAPP"."""
    content_bytes = _build_xlsx([{"case_platform": "PC端, , WJAPP端"}])
    result = _read_with_enum(content_bytes, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert platform_errs == []
    assert result.valid_cases[0]["case_platform"] == "PC,WJAPP"


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m1_only_separators_is_required_error():
    """M1 解析: cell = ",,, " 全部空段 -> 视为空值, 报 "不能为空"."""
    content_bytes = _build_xlsx([{"case_platform": ",,, "}])
    result = _read_with_enum(content_bytes, VALID_ENUM)
    flat = [item for row_err in result.errors for item in row_err.get("errors", [])]
    platform_errs = [e for e in flat if e.get("field") == "适用端"]
    assert len(platform_errs) == 1
    assert "不能为空" in platform_errs[0]["message"]


# ----- M2 commit 路径多值测试 -----

@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_multi_platform_validated_and_converted(patch_enum_config):
    """M2 commit 校验: cell = "PC端, WJAPP端" -> 原地改 value 为 "PC,WJAPP"."""
    patch_enum_config(PLATFORM_MAP)
    cases = [{"case_platform": "PC端, WJAPP端", "_row": 2}]
    await M2ImportService._validate_field_constraints(cases)
    assert cases[0]["case_platform"] == "PC,WJAPP"


@pytest.mark.unit
@pytest.mark.asyncio
async def test_m2_multi_platform_invalid_part_raises(patch_enum_config):
    """M2 commit 校验: cell 含非法 label -> 整条 raise, 列出非法 labels."""
    patch_enum_config(PLATFORM_MAP)
    cases = [{"case_platform": "PC端, XBOX端", "_row": 3}]
    with pytest.raises(CommonError) as exc_info:
        await M2ImportService._validate_field_constraints(cases)
    err = exc_info.value
    assert "XBOX端" in str(err.message)
    # 非法 label 时不应改写原值
    assert cases[0]["case_platform"] == "PC端, XBOX端"
    # 错误数据带 row 上下文 (跟 controller 行号错误对齐)
    assert err.data["errors"][0]["row"] == 3


# ----- helper 函数纯单元测试 -----

@pytest.mark.unit
def test_helpers_split_and_join_roundtrip():
    from utils.caseEnumResolver import split_platform_value, join_platform_value
    # split 跟 join 互逆
    assert join_platform_value(split_platform_value("PC,WJAPP,APP")) == "PC,WJAPP,APP"
    assert split_platform_value("") == []
    assert split_platform_value(None) == []
    # 各种 trim / 去重 / 去空
    assert split_platform_value("PC, PC, ,PC,") == ["PC"]
    assert split_platform_value(",PC,WJAPP,") == ["PC", "WJAPP"]


@pytest.mark.unit
def test_helper_format_platform_value_for_export_multi():
    from utils.caseEnumResolver import format_platform_value_for_export
    label_map = {"PC": "PC端", "WJAPP": "WJAPP端", "APP": "APP端"}
    assert format_platform_value_for_export("PC,WJAPP", label_map) == "PC端,WJAPP端"
    assert format_platform_value_for_export("PC", label_map) == "PC端"
    assert format_platform_value_for_export(None, label_map) is None
    # 缺 label 时回退原 value, 跟单值 label_map.get(v, v) 语义一致
    assert format_platform_value_for_export("PC,XBOX", label_map) == "PC端,XBOX"


@pytest.mark.unit
def test_helper_parse_platform_cell_dedup_and_order():
    from utils.caseEnumResolver import parse_platform_cell_to_value
    platform_map = PLATFORM_MAP
    v, err = parse_platform_cell_to_value("PC端, WJAPP端, PC端", platform_map)
    assert v == "PC,WJAPP" and err is None
    v, err = parse_platform_cell_to_value(" WJAPP端 , PC端 ", platform_map)
    assert v == "WJAPP,PC" and err is None
    # cell 是 float NaN (pandas read 出来常见)
    import math
    v, err = parse_platform_cell_to_value(float("nan"), platform_map)
    assert v is None and err is None
    # cell 是 None
    v, err = parse_platform_cell_to_value(None, platform_map)
    assert v is None and err is None


# --------------------------------------------------------------------------- #
# 导出端: DataValidation errorStyle="warning" + F2 comment
# --------------------------------------------------------------------------- #

@pytest.mark.unit
def test_export_platform_dv_uses_stop_error_style():
    """
    适用端 (F 列) DataValidation 用默认 stop (不写 errorStyle). 因为下拉里
    已经预生成所有 2^n - 1 个单/多值组合, 用户从下拉里选就行, 没必要 warning
    弹窗让用户手输. 真要手输错值, stop 反而能阻止脏数据.
    """
    from app.service.exportCaseService import ExportCaseService
    import os, tempfile, zipfile

    service = ExportCaseService(
        scope_type="library",
        scope_id=1,
        case_dicts=[{"id": 1, "case_name": "demo", "case_platform": "PC,WJAPP"}],
        group_path_map={1: "默认分组"},
        label_map={"PC": "PC端", "WJAPP": "WJAPP端"},
        all_group_paths=["默认分组"],
        level_labels=["P0"],
        type_labels=["功能"],
        # 模拟 controller 已经把 combo 展开成 7 个选项 (3 个 platform labels)
        platform_labels=["AA", "BB", "CC", "AA,BB", "AA,CC", "BB,CC", "AA,BB,CC"],
    )
    buf = service.build_workbook()
    buf.seek(0)
    with tempfile.TemporaryDirectory() as td:
        xlsx_path = os.path.join(td, "out.xlsx")
        with open(xlsx_path, "wb") as f:
            f.write(buf.getvalue())
        with zipfile.ZipFile(xlsx_path) as z:
            sheet_xml = z.read("xl/worksheets/sheet1.xml").decode("utf-8")
    # F 列 dv 引用 _dv!C 列
    assert "=_dv!$C$1:$C$" in sheet_xml
    # 不应是 warning (旧实现用过, 已回退)
    assert "errorStyle" not in sheet_xml, (
        "F 列 dv 不应写 errorStyle, 走默认 stop 即可 (下拉里已含所有组合)"
    )


@pytest.mark.unit
def test_export_platform_header_cell_has_no_comment():
    """F2 标题 cell 不再有 comment. 提示已经写到 DataValidation prompt 字段, 冗余 comment 反而是噪音."""
    from app.service.exportCaseService import ExportCaseService
    import openpyxl

    service = ExportCaseService(
        scope_type="library",
        scope_id=1,
        case_dicts=[{"id": 1, "case_name": "demo", "case_platform": "PC"}],
        group_path_map={1: "默认分组"},
        label_map={"PC": "PC端"},
        all_group_paths=["默认分组"],
        level_labels=["P0"],
        type_labels=["功能"],
        platform_labels=["PC端"],
    )
    buf = service.build_workbook()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb["用例数据"]
    cell = ws.cell(row=2, column=6)  # F2 = "适用端" header
    assert cell.value == "适用端"
    assert cell.comment is None, "F2 不应有 comment (提示在 dv prompt 字段)"


@pytest.mark.unit
def test_export_platform_value_rendered_as_labels_for_multi_value():
    """
    导出: case_platform="PC,WJAPP" -> cell 显示 "PC端,WJAPP端".
    验证多值在导出端被正确翻成 label 串, 不留 value.
    """
    from app.service.exportCaseService import ExportCaseService
    import openpyxl

    service = ExportCaseService(
        scope_type="library",
        scope_id=1,
        case_dicts=[{"id": 1, "case_name": "demo", "case_platform": "PC,WJAPP"}],
        group_path_map={1: "默认分组"},
        label_map={"PC": "PC端", "WJAPP": "WJAPP端"},
        all_group_paths=["默认分组"],
        level_labels=["P0"],
        type_labels=["功能"],
        platform_labels=["PC端", "WJAPP端"],
    )
    buf = service.build_workbook()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    ws = wb["用例数据"]
    # F3 是第一行数据
    assert ws.cell(row=3, column=6).value == "PC端,WJAPP端"


# --------------------------------------------------------------------------- #
# 适用端 combo label 生成器 (dropdown 内容源)
# --------------------------------------------------------------------------- #

@pytest.mark.unit
def test_combo_n1_single_label():
    """n=1: 只有 1 个 label, combo 也只有 1 个."""
    from utils.caseEnumResolver import generate_platform_combo_labels
    assert generate_platform_combo_labels({"AA": "AA"}) == ["AA"]


@pytest.mark.unit
def test_combo_n3_full_expansion():
    """n=3: AA/BB/CC -> 7 个组合, 长度升序, 同长度字典序."""
    from utils.caseEnumResolver import generate_platform_combo_labels
    result = generate_platform_combo_labels({"AA": "AA", "BB": "BB", "CC": "CC"})
    assert result == [
        "AA", "BB", "CC",
        "AA,BB", "AA,CC", "BB,CC",
        "AA,BB,CC",
    ]


@pytest.mark.unit
def test_combo_n2_pair():
    """n=2: AA/BB -> 3 个组合."""
    from utils.caseEnumResolver import generate_platform_combo_labels
    assert generate_platform_combo_labels({"AA": "AA", "BB": "BB"}) == [
        "AA", "BB", "AA,BB",
    ]


@pytest.mark.unit
def test_combo_n4_eight_options():
    """n=4: 2^4 - 1 = 15 个组合."""
    from itertools import combinations
    from utils.caseEnumResolver import generate_platform_combo_labels
    labels = ["A", "B", "C", "D"]
    result = generate_platform_combo_labels({k: k for k in labels})
    assert len(result) == 15
    # 验证单值全部在最前
    assert result[:4] == ["A", "B", "C", "D"]
    # 验证全集在最后
    assert result[-1] == "A,B,C,D"
    # 验证长度升序
    for i in range(len(result) - 1):
        a, b = result[i].count(","), result[i + 1].count(",")
        assert a <= b, f"长度应升序: {result[i]!r} ({a}) >= {result[i + 1]!r} ({b})"


@pytest.mark.unit
def test_combo_empty_map_returns_empty():
    """空 map: 返回空 list, 下拉跳过该列."""
    from utils.caseEnumResolver import generate_platform_combo_labels
    assert generate_platform_combo_labels({}) == []


@pytest.mark.unit
def test_combo_label_value_distinct():
    """label 跟 value 可以不同, combo 仍按 label 拼 (用户看到的字符)."""
    from utils.caseEnumResolver import generate_platform_combo_labels
    # 实际业务: label="PC端" value="PC"
    result = generate_platform_combo_labels({
        "PC端": "PC", "H5端": "H5",
    })
    assert result == ["H5端", "PC端", "H5端,PC端"]
    # value 不会出现在 combo 串里 (那是 DB 存的内部表示)


@pytest.mark.unit
def test_combo_too_many_labels_falls_back_to_single():
    """n > PLATFORM_COMBO_MAX_LABELS: 退化到只返单值, 不爆 dropdown."""
    from utils.caseEnumResolver import generate_platform_combo_labels
    big = {f"P{i}端": f"P{i}" for i in range(10)}  # n=10, 2^10 - 1 = 1023
    result = generate_platform_combo_labels(big)
    # 应该只返 10 个单值 (按 label 排序)
    assert len(result) == 10
    assert result == sorted(big.keys())


# --------------------------------------------------------------------------- #
# End-to-end: load_case_enum_label_lists 真的把 PLATFORM 展开成 combo
# --------------------------------------------------------------------------- #

@pytest.mark.unit
@pytest.mark.asyncio
async def test_load_case_enum_label_lists_platform_uses_combo(monkeypatch):
    """
    load_case_enum_label_lists 的 PLATFORM key 必须走 combo 展开, 不能只返单值.
    模拟 CaseConfigMapper 返回 3 个 platform config, 验证输出有 7 个选项.
    """
    from utils.caseEnumResolver import load_case_enum_label_lists, PLATFORM_CONFIG_KEY

    class FakeCfg:
        def __init__(self, label, value):
            self.label = label
            self.value = value
            self.enabled = 1

    async def fake_query(key):
        if key == "PLATFORM":
            return [FakeCfg("AA", "AA"), FakeCfg("BB", "BB"), FakeCfg("CC", "CC")]
        return []

    monkeypatch.setattr(
        "app.mapper.test_case.caseConfigMapper.CaseConfigMapper.query_by_key",
        fake_query,
    )
    result = await load_case_enum_label_lists()
    assert PLATFORM_CONFIG_KEY in result
    assert result[PLATFORM_CONFIG_KEY] == [
        "AA", "BB", "CC", "AA,BB", "AA,CC", "BB,CC", "AA,BB,CC",
    ]


@pytest.mark.unit
@pytest.mark.asyncio
async def test_load_case_enum_label_lists_others_unchanged(monkeypatch):
    """
    CASE_LEVEL / CASE_TYPE 不走 combo, 保持原样 (单值列表). 跟 PLATFORM 区别对待.
    """
    from utils.caseEnumResolver import load_case_enum_label_lists

    class FakeCfg:
        def __init__(self, label, value):
            self.label = label
            self.value = value
            self.enabled = 1

    async def fake_query(key):
        return {
            "CASE_LEVEL": [FakeCfg("P0", "P0"), FakeCfg("P1", "P1")],
            "CASE_TYPE": [FakeCfg("功能", "GN"), FakeCfg("冒烟", "MY")],
            "PLATFORM": [FakeCfg("PC", "PC"), FakeCfg("H5", "H5")],
        }.get(key, [])

    monkeypatch.setattr(
        "app.mapper.test_case.caseConfigMapper.CaseConfigMapper.query_by_key",
        fake_query,
    )
    result = await load_case_enum_label_lists()
    # LEVEL/TYPE 不展开 combo
    assert result["CASE_LEVEL"] == ["P0", "P1"]
    assert result["CASE_TYPE"] == ["功能", "冒烟"]
    # PLATFORM 展开
    assert result["PLATFORM"] == ["H5", "PC", "H5,PC"]


# --------------------------------------------------------------------------- #
# End-to-end: 导出 dropdown 真的把 combo 写进 _dv sheet
# --------------------------------------------------------------------------- #

@pytest.mark.unit
def test_export_dv_sheet_contains_all_combos():
    """
    导出时 _dv sheet 的 C 列 (适用端下拉源) 必须包含所有 2^n - 1 个 combo.
    controller 负责在传 platform_labels 之前先调 load_case_enum_label_lists,
    这里直接模拟传 combo 列表进来, 验证 _dv sheet 真的把它原样落库.
    """
    from app.service.exportCaseService import ExportCaseService
    import openpyxl

    # 模拟 controller 已调过 load_case_enum_label_lists 拿到 7 个 combo
    combos = ["AA", "BB", "CC", "AA,BB", "AA,CC", "BB,CC", "AA,BB,CC"]
    service = ExportCaseService(
        scope_type="library",
        scope_id=1,
        case_dicts=[{"id": 1, "case_name": "demo", "case_platform": "AA,BB"}],
        group_path_map={1: "默认分组"},
        label_map={"AA": "AA", "BB": "BB", "CC": "CC"},
        all_group_paths=["默认分组"],
        level_labels=["P0"],
        type_labels=["功能"],
        platform_labels=combos,
    )
    buf = service.build_workbook()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf)
    dv_ws = wb["_dv"]
    # C 列 (适用端) 从 row 1 开始连续写 combo
    written = []
    for row in dv_ws.iter_rows(min_col=3, max_col=3, values_only=True):
        if row[0]:
            written.append(row[0])
        else:
            break
    assert written == combos, (
        f"_dv!C 列应包含所有 7 个 combo, 实际: {written}"
    )
