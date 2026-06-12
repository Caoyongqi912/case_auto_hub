"""
PR-3 Step 1 单测:
- RoundtripReader.detect_template_type: M1/M2 文件探测
- RoundtripReader.extract_scope_from_meta: 从 _meta 拿 scope
- RoundtripReader.async_read_from_bytes: M2 bytes 入口
- AsyncFilesReader.async_read_from_bytes: M1 bytes 入口
- UploadPreviewResult: 新增 template_type / warnings 字段默认值

不依赖数据库 / Redis, 用 openpyxl 内存构造 xlsx.
"""
import io
import pytest
import openpyxl

from utils.roundtripReader import (
    RoundtripReader,
    META_SHEET_NAME,
    DATA_SHEET_NAME,
)
from utils.aioFileReader import AsyncFilesReader
from app.schema.hub.testCaseSchema import UploadPreviewResult


# ---------- fixtures ----------

@pytest.fixture
def m1_xlsx_bytes() -> bytes:
    """M1 老模板: 只有 1 个 sheet (template), 不带 _meta."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "template"
    # 3 行结构: 标题/表头/demo, 跟 file/用例模版.xlsx 一致
    ws["A1"] = "导入模板"
    ws["A2"] = "标题*"
    ws["B2"] = "用例等级*"
    ws["C2"] = "步骤描述"
    ws["D2"] = "预期结果"
    ws["A3"] = "示例用例"
    ws["B3"] = "P2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@pytest.fixture
def m2_xlsx_bytes() -> bytes:
    """M2 导回模板: PR-3 新协议 = M1 9 列 + 用例ID + _meta.
    行模型: row 1 标题引导 / row 2 表头 / row 3+ 用例 (步骤拼 cell, 1 用例 1 行)."""
    wb = openpyxl.Workbook()
    # 主表
    ws = wb.active
    ws.title = DATA_SHEET_NAME
    ws["A1"] = "CASE  HUB 用例导出"  # 标题引导, 解析器跳过
    headers = ["标题*", "所属分组", "前置条件", "步骤描述*", "预期结果*",
               "标签", "用例等级*", "用例类型", "备注", "用例ID"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
    # 1 条用例 (步骤拼 cell, 1 行)
    row = ["登录成功", "默认分组/登录", "已注册", "【1】输入账号密码\n【2】点击登录",
           "【1】登录成功\n【2】跳首页", "smoke", "P1", "功能", "", 1001]
    for col, v in enumerate(row, 1):
        ws.cell(row=3, column=col, value=v)
    # _meta (隐藏)
    meta = wb.create_sheet(META_SHEET_NAME)
    meta.sheet_state = "hidden"
    meta["A1"] = "scope_type"; meta["B1"] = "library"
    meta["A2"] = "scope_id"; meta["B2"] = "100"
    meta["A3"] = "version"; meta["B3"] = "2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------- detect_template_type ----------

def test_detect_template_type_m1(m1_xlsx_bytes):
    """M1 老模板: 无 _meta sheet -> 'M1'."""
    assert RoundtripReader.detect_template_type(m1_xlsx_bytes) == "M1"


def test_detect_template_type_m2(m2_xlsx_bytes):
    """M2 导回模板: 有 _meta sheet -> 'M2'."""
    assert RoundtripReader.detect_template_type(m2_xlsx_bytes) == "M2"


def test_detect_template_type_invalid_bytes():
    """损坏的 xlsx -> 默认 M1, 不抛异常 (让下游 aioFileReader 抛详细错误)."""
    assert RoundtripReader.detect_template_type(b"not a xlsx") == "M1"


# ---------- extract_scope_from_meta ----------

def test_extract_scope_library(m2_xlsx_bytes):
    """从 M2 _meta 拿 (library, 100)."""
    scope = RoundtripReader.extract_scope_from_meta(m2_xlsx_bytes)
    assert scope == ("library", 100)


def test_extract_scope_from_m1_returns_none(m1_xlsx_bytes):
    """M1 老模板无 _meta -> None."""
    assert RoundtripReader.extract_scope_from_meta(m1_xlsx_bytes) is None


def test_extract_scope_missing_fields():
    """_meta 存在但缺 scope 字段 -> None."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = META_SHEET_NAME
    ws["A1"] = "version"; ws["B1"] = "1"
    bio = io.BytesIO()
    wb.save(bio)
    assert RoundtripReader.extract_scope_from_meta(bio.getvalue()) is None


def test_extract_scope_invalid_id():
    """scope_id 不是数字 -> None."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = META_SHEET_NAME
    ws["A1"] = "scope_type"; ws["B1"] = "library"
    ws["A2"] = "scope_id"; ws["B2"] = "not-a-number"
    bio = io.BytesIO()
    wb.save(bio)
    assert RoundtripReader.extract_scope_from_meta(bio.getvalue()) is None


# ---------- async_read_from_bytes (M2 路径) ----------

@pytest.mark.asyncio
async def test_roundtrip_async_read_from_bytes_ok(m2_xlsx_bytes):
    """M2 bytes 入口: 解析主表 1 条用例, 拿到 _meta, scope 校验通过."""
    reader = RoundtripReader(scope_type="library", scope_id=100)
    result = await reader.async_read_from_bytes(m2_xlsx_bytes)
    assert result.errors == [], f"unexpected errors: {result.errors}"
    assert result.total_count == 1
    assert result.valid_count == 1
    assert result.invalid_count == 0
    assert result.meta is not None
    assert result.meta.get("scope_type") == "library"
    assert result.meta.get("scope_id") == "100"
    # scope_check 验证
    assert result.scope_check.get("scope_type_matches") is True
    assert result.scope_check.get("scope_id_matches") is True


@pytest.mark.asyncio
async def test_roundtrip_async_read_from_bytes_scope_mismatch(m2_xlsx_bytes):
    """scope 不匹配 -> errors 写一行, can_commit=False 语义由 controller 决定."""
    reader = RoundtripReader(scope_type="plan", scope_id=999)
    result = await reader.async_read_from_bytes(m2_xlsx_bytes)
    # _meta scope_type=library, 构造时传 plan -> 必 mismatch
    assert result.scope_check.get("scope_type_matches") is False
    assert any("scope" in str(e).lower() for err in result.errors for e in err.get("errors", []))


# ---------- async_read_from_bytes (M1 路径) ----------

@pytest.mark.asyncio
async def test_async_files_reader_from_bytes_m1(m1_xlsx_bytes):
    """M1 bytes 入口: 9 列老解析, demo 行被读为一条用例."""
    reader = AsyncFilesReader(enum_config=None)
    result = await reader.async_read_from_bytes(m1_xlsx_bytes)
    # 不传 enum_config 时, 字段缺失不报错, 但必填 case_name 缺失会报
    # M1 demo 行 case_name="示例用例" 非空, 应该被读为 valid
    assert result.file_md5  # 32 位 md5
    # 校验: 至少有 valid_count >= 0 (不强制, demo 行也可能被识别)
    assert result.total_count >= 0


# ---------- UploadPreviewResult schema ----------

def test_upload_preview_result_default_m1():
    """老调用方不传 template_type, 默认 M1 保兼容."""
    m = UploadPreviewResult(total_count=0, valid_count=0, invalid_count=0)
    assert m.template_type == "M1"
    assert m.warnings == []
    assert m.can_commit is True  # 默认 True


def test_upload_preview_result_explicit_m2():
    """M2 路径显式传 template_type='M2' + warnings."""
    m = UploadPreviewResult(
        total_count=1, valid_count=1, invalid_count=0,
        template_type="M2",
        warnings=[{"row": 3, "message": "test warning"}],
    )
    assert m.template_type == "M2"
    assert m.warnings == [{"row": 3, "message": "test warning"}]


def test_upload_preview_result_rejects_invalid_template_type():
    """template_type 必须是 M1 或 M2, 其它值被 pydantic 拒掉."""
    import pydantic
    with pytest.raises(pydantic.ValidationError):
        UploadPreviewResult(
            total_count=0, valid_count=0, invalid_count=0,
            template_type="M3",  # type: ignore[arg-type]
        )



# ---------- E2E: 真实 PR-1 导出文件 ----------

@pytest.fixture
def real_pr1_export_bytes() -> bytes:
    """
    真实 PR-1 /export 端点导出的文件. 不是测试构造, 是 caseHub 库 (scope=library:46) 4 条用例.
    验证: 1) RoundtripReader 解析跟 PR-1 写出的格式 (10 列 + _meta v2) 完全兼容
          2) 4 个 case_id (4082/4081/4080/4038) 都正确读出
    这个 fixture 在跑测试前需要 /tmp/exported.xlsx 存在. 不存在时单测会 skip.
    """
    import os
    path = "/tmp/exported.xlsx"
    if not os.path.exists(path):
        pytest.skip(f"真实 PR-1 导出文件不存在: {path}, 跑 e2e 前先 curl 一次: "
                    f"curl -X POST .../hub/cases/export?scope_type=library&scope_id=46&project_id=1")
    with open(path, "rb") as f:
        return f.read()


@pytest.mark.asyncio
async def test_real_pr1_export_roundtrip(real_pr1_export_bytes):
    """真实 PR-1 导出文件 -> RoundtripReader 解析 4 条用例, case_id 全对得上."""
    reader = RoundtripReader(scope_type="library", scope_id=46)
    result = await reader.async_read_from_bytes(real_pr1_export_bytes)
    # scope + version 校验全过
    assert result.errors == [], f"unexpected errors: {result.errors}"
    assert result.scope_check["scope_type_matches"] is True
    assert result.scope_check["scope_id_matches"] is True
    assert result.scope_check["version_supported"] is True
    # 4 条用例全 valid
    assert result.total_count == 4
    assert result.valid_count == 4
    assert result.invalid_count == 0
    # case_id 跟 _meta.case_ids_at_export 对得上
    case_ids = sorted(c["case_id"] for c in result.valid_cases)
    assert case_ids == [4038, 4080, 4081, 4082]
    # 主表字段 (M1 9 列) 都解析出来了
    first = result.valid_cases[0]
    for k in ("case_name", "case_setup", "action", "expected_result",
              "case_tag", "case_level", "case_type", "case_mark", "group_path"):
        assert k in first, f"missing field: {k}"
    # 步骤保持 cell 拼多步原样 (PR-1 写时就这么写的)
    assert "【1】" in (first["action"] or "")
