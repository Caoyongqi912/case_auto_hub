#!/usr/bin/env python
# -*- coding:utf-8 -*-
# @Time : 2026/6/12
# @Author : cyq
# @File : roundtripReader
# @Software: PyCharm
# @Desc: 导出-编辑-导回 导回专用的 Excel 解析 (PR-3 重写版)
#
# 新协议 (跟之前 PR-2 写的 14 列不一样, 按用户最新 plan):
#   M2 文件 = M1 9 列模板 + "用例ID" 列 + 隐藏 _meta Sheet
#   - 主表 10 列: 标题* / 所属分组 / 前置条件 / 步骤描述* / 预期结果* / 标签 / 用例等级* / 用例类型 / 备注 / 用例ID
#   - _meta: scope_type / scope_id / case_ids_at_export / exported_at / version=2
#
# 跟 utils/aioFileReader.py 的关系:
#   - aioFileReader 解析 M1 老 9 列 (FIELD_MAPPING), 本解析器复用它解析 M1 9 列
#   - 步骤 (步骤描述 / 预期结果) 保持 cell 拼多步原样 (跟 PR-1 写出一致), 不做 1 步 1 行拆分
#   - 多读 1 列 "用例ID" (FIELD_MAPPING 已加) + 单独读 _meta
#
# 输入文件结构 (PR-1 export 出的):
#   Sheet 1  "用例数据"  — 10 列主表, 步骤拼 cell ("【1】xxx\\n【2】xxx")
#   Sheet 2  "编辑指引"  — 可见, 解析器跳过
#   Sheet 3  "_meta"     — 隐藏, scope_type / scope_id / case_ids_at_export / version
import asyncio
import hashlib
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any, Dict, List, Literal, Optional, Tuple

import openpyxl
import pandas as pd
from fastapi import UploadFile
from utils import log
from utils.aioFileReader import AsyncFilesReader, CaseEnumConfig
from utils.threadPool import ThreadPoolHelper

# 主表 Sheet 名 (跟 PR-1 export 写出一致)
DATA_SHEET_NAME = "用例数据"
META_SHEET_NAME = "_meta"

# 协议版本. _meta.version 不一致视为不兼容.
# 跟 exportCaseService.META_VERSION=2 对齐 (PR-1 写 v2, 本解析端读 v2).
SUPPORTED_META_VERSION = 2
MAX_FILE_SIZE = 10 * 1024 * 1024


@dataclass
class RoundtripParseResult:
    file_md5: str = ""
    total_count: int = 0
    valid_count: int = 0
    invalid_count: int = 0
    errors: List[Dict[str, Any]] = field(default_factory=list)
    # 每行解析后的 dict, key 跟 aioFileReader 一致 (case_name / case_setup / action /
    # expected_result / case_tag / case_level / case_type / case_mark / group_path)
    # + case_id (M2 独有) + _row (Excel 物理行号, commit 前 controller 剥掉)
    valid_cases: List[Dict[str, Any]] = field(default_factory=list)
    # _meta Sheet 解析结果. 缺失或字段缺失时为 None
    meta: Optional[Dict[str, str]] = None
    # scope 校验结果. 失败时 errors 也会带一项.
    scope_check: Dict[str, Any] = field(default_factory=dict)


class RoundtripReader:
    """
    解析 10 列 xlsx (PR-1 导出格式) -> RoundtripParseResult.

    流程:
      1) 复用 AsyncFilesReader 解析 M1 9 列 + "用例ID" 列 (FIELD_MAPPING 已有)
      2) 单独读 _meta Sheet (openpyxl)
      3) scope 校验 (form scope_type/scope_id vs _meta)
      4) 错误时 errors 里带, can_commit=false (controller 决定)
    """

    def __init__(self, scope_type: str, scope_id: int, workers: int = 4):
        if scope_type not in ("library", "plan"):
            raise ValueError(f"scope_type 必须是 library 或 plan, 收到: {scope_type!r}")
        self.scope_type = scope_type
        self.scope_id = scope_id
        self.tph = ThreadPoolHelper(workers=workers)

    async def async_read(self, file: UploadFile) -> RoundtripParseResult:
        content = await file.read()
        return await self.async_read_from_bytes(content)

    async def async_read_from_bytes(self, content: bytes) -> RoundtripParseResult:
        if len(content) > MAX_FILE_SIZE:
            raise ValueError(f"文件大小超过限制, 最大支持 {MAX_FILE_SIZE // 1024 // 1024}MB")
        file_bytes = BytesIO(content)
        file_md5 = hashlib.md5(content).hexdigest()
        loop = asyncio.get_event_loop()
        return await self.tph.run_in_exe(loop, self._read, file_bytes, file_md5)

    def _read(self, file: BytesIO, file_md5: str) -> RoundtripParseResult:
        result = RoundtripParseResult(file_md5=file_md5)

        # 1) 主表 sheet 必须存在
        try:
            excel = pd.ExcelFile(file, engine="openpyxl")
            sheet_names = excel.sheet_names
        except Exception as e:
            log.error(f"roundtrip 读盘失败: {e}")
            raise ValueError(f"读取 Excel 时发生错误: {e}")
        if DATA_SHEET_NAME not in sheet_names:
            result.errors.append({
                "row": 0,
                "errors": [{
                    "field": "file",
                    "message": f"找不到主表 Sheet '{DATA_SHEET_NAME}', 不是导回格式",
                }],
            })
            return result

        # 2) 解析 _meta (隐藏 sheet). 缺 _meta 直接打回 (M2 必备)
        result.meta = self._read_meta(excel)
        if result.meta is None:
            result.errors.append({
                "row": 0,
                "errors": [{
                    "field": "_meta",
                    "message": "找不到 _meta Sheet, 不是导回格式",
                }],
            })
            return result

        # 3) scope 校验. 不一致提前失败, 不浪费主表解析
        self._validate_meta(result)
        if result.scope_check.get("scope_type_matches") is False:
            result.errors.append({
                "row": 0,
                "errors": [{
                    "field": "scope_type",
                    "message": (
                        f"文件导出于 {result.meta.get('scope_type')}:{result.meta.get('scope_id')}, "
                        f"与当前选择的 {self.scope_type}:{self.scope_id} 不一致"
                    ),
                }],
            })
            return result
        if result.scope_check.get("scope_id_matches") is False:
            result.errors.append({
                "row": 0,
                "errors": [{
                    "field": "scope_id",
                    "message": (
                        f"文件导出于 {result.meta.get('scope_type')}:{result.meta.get('scope_id')}, "
                        f"与当前选择的 {self.scope_type}:{self.scope_id} 不一致"
                    ),
                }],
            })
            return result
        if result.scope_check.get("version_supported") is False:
            result.errors.append({
                "row": 0,
                "errors": [{
                    "field": "version",
                    "message": f"文件版本 {result.meta.get('version')} 不受支持, 期望 {SUPPORTED_META_VERSION}",
                }],
            })
            return result

        # 4) 主表解析: 复用 AsyncFilesReader 解析 M1 9 列 + "用例ID" 列.
        # AsyncFilesReader 接受 FILE-like (BytesIO), 我们传个新 BytesIO 因为上面 _read_meta
        # 把 file 指针用过了. file 是 BytesIO, 可以 seek(0) 重置.
        file.seek(0)
        # enum_config=None 时不做 label->value 转换, 留原 label. M2 解析阶段不卡枚举 (commit 再校验)
        aio_result = AsyncFilesReader(enum_config=None)._read(  # noqa: SLF001
            file=file, file_md5=file_md5, enum_config=CaseEnumConfig(),
        )

        # 5) 主表解析结果: 把 _row 留着 (跟 aioFileReader 一致), case_id 已经塞进 mapped_case
        result.valid_cases = aio_result.valid_cases
        result.errors.extend(aio_result.errors)
        result.total_count = aio_result.total_count
        result.valid_count = len(aio_result.valid_cases)
        result.invalid_count = len(result.errors)

        return result

    def _read_meta(self, excel: pd.ExcelFile) -> Optional[Dict[str, str]]:
        if META_SHEET_NAME not in excel.sheet_names:
            return None
        try:
            df = pd.read_excel(
                excel, sheet_name=META_SHEET_NAME, header=None,
                engine="openpyxl", keep_default_na=False,
            )
        except Exception as e:
            log.warning(f"读 _meta 失败: {e}")
            return None
        meta: Dict[str, str] = {}
        for _, row in df.iterrows():
            if len(row) < 2:
                continue
            k = str(row.iloc[0]).strip() if row.iloc[0] is not None else ""
            v = str(row.iloc[1]).strip() if row.iloc[1] is not None else ""
            if k:
                meta[k] = v
        return meta or None

    def _validate_meta(self, result: RoundtripParseResult) -> None:
        meta = result.meta or {}
        file_scope_type = meta.get("scope_type")
        file_scope_id = meta.get("scope_id")
        version = meta.get("version")

        result.scope_check = {
            "scope_type_matches": file_scope_type == self.scope_type,
            "scope_id_matches": str(file_scope_id) == str(self.scope_id) if file_scope_id is not None else False,
            "version_supported": (int(version) == SUPPORTED_META_VERSION) if version and version.lstrip("-").isdigit() else False,
            "meta_scope_type": file_scope_type,
            "meta_scope_id": file_scope_id,
        }

    @staticmethod
    def detect_template_type(content: bytes) -> Literal["M1", "M2"]:
        """
        探测 xlsx 模板类型: 看是否含 _meta Sheet.
        - 有 _meta -> M2 (导回格式)
        - 无 _meta -> M1 (老 9 列导入模板)
        探测失败时返回 M1, 让下游 aioFileReader 抛详细错误.
        """
        try:
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            is_m2 = META_SHEET_NAME in wb.sheetnames
            wb.close()
            return "M2" if is_m2 else "M1"
        except Exception as e:
            log.warning(f"detect_template_type 失败, 默认 M1: {e}")
            return "M1"

    @staticmethod
    def extract_scope_from_meta(content: bytes) -> Optional[Tuple[str, int]]:
        """
        从 _meta Sheet 读 (scope_type, scope_id). 给 /upload 入口探测到 M2 后
        用来构造 RoundtripReader. _meta 缺失 / scope 字段缺失 / scope_type
        不是 library|plan / scope_id 不是数字 都返回 None, 此时 controller
        用占位 ("library", 0), 让下游 _validate_meta 把 scope 错误写入 errors.
        """
        try:
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            if META_SHEET_NAME not in wb.sheetnames:
                wb.close()
                return None
            ws = wb[META_SHEET_NAME]
            meta: Dict[str, str] = {}
            for row in ws.iter_rows(values_only=True):
                if not row or len(row) < 2:
                    continue
                k = row[0]
                v = row[1]
                if k is None:
                    continue
                meta[str(k).strip()] = "" if v is None else str(v).strip()
            wb.close()
            st = meta.get("scope_type")
            sid = meta.get("scope_id")
            if st in ("library", "plan") and sid and sid.lstrip("-").isdigit():
                return st, int(sid)
        except Exception as e:
            log.warning(f"extract_scope_from_meta 失败: {e}")
        return None
