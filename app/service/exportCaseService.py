"""
用例导出服务: 基于 file/用例模版.xlsx 模板生成导出文件.

复用上传模板的 3 行结构 (标题引导 + 标头 + demo) 保证视觉一致, 增量写入:
  - 第 10 列 "用例ID" (下载模板没有, 导出才有)
  - 第 4 行起的真实数据 (下载模板无数据)
  - 隐藏的 _meta Sheet (跨 scope 防御 / 变更检测 / 版本号)

行布局 (基于 file/用例模版.xlsx 复用样式, 不是逐行复制):
  Row 1   - A1 单元格里嵌标题 + 编辑引导 (覆盖原 "导入模板" 标题)
  Row 2   - 10 列表头 (前 9 列沿用模板, 第 10 列 "用例ID" 新增)
  Row 3+  - 真实数据 (10 列, 1 用例 1 行, 多步拼到同一 cell).
             原模板 row 3 的 demo 行被删除, 避免再导入时被当成 INSERT 假数据.
"""

from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.model.base.module import Module
from app.model.caseHub.plan_module import PlanModule

# 导出 Excel 下拉框枚举. 业务上**不**写死:
# - 等级/类型/适用端的下拉项均由 controller 从 case_enum_config 加载后传入 ExportCaseService,
#   反映 admin 在配置中心的最新增删.
# - 这里的 EXPORT_LEVEL_OPTIONS / EXPORT_TYPE_OPTIONS / EXPORT_PLATFORM_OPTIONS 仅作
#   controller 拉取失败时的兜底空列表 (Excel 下拉拿空列表就跳过该列 dv, 不报错).
# 历史保留这段注释, 提醒不要在 service 里再写死业务枚举.
# EXPORT_LEVEL_OPTIONS = ["P0-最高", "P1-高", "P2-中", "P3-低"]   # 已废弃
# EXPORT_TYPE_OPTIONS = ["功能测试", "冒烟", "回归", "其他"]      # 已废弃
# EXPORT_PLATFORM_OPTIONS = ["PC", "H5"]                          # 已废弃

# 单次导出硬上限. 超量走二期异步任务.
EXPORT_HARD_LIMIT = 10_000

# 导出文件协议版本. 与下载模板 (version=N/A) 区分, 解析端可据此拒绝老格式.
META_VERSION = 2

# 导出 Sheet 名. 模板里原叫 "template", 导出改名让用户更直观.
SHEET_DATA = "用例数据"
SHEET_META = "_meta"
# 下拉源 Sheet. 所有 DataValidation 引用都走这里, 主表不再写 Z/AA/AB/AC 隐藏列.
# 跨 sheet 引用是 OOXML 标准, WPS / Excel / LibreOffice 全兼容.
# 写在这里的好处: 主表彻底干净, 不会因为用户手填/复制把隐藏列污染,
# 也不会让 _is_empty_row 误判空行.
SHEET_DV = "_dv"
# _dv Sheet 内列布局: A=用例等级, B=用例类型, C=适用端, D=所属分组
_DV_LEVEL_COL = 1     # A
_DV_TYPE_COL = 2      # B
_DV_PLATFORM_COL = 3  # C
_DV_GROUP_COL = 4     # D

# 编辑引导. 行 1 的 A1 cell 全文, 包含标题 + 5 条规则. 规则从用户立场写, 不说后端实现.
_EXPORT_GUIDE = """\
CASE  HUB 用例导出

1. 数据列 (标题 / 前置条件 / 步骤描述 / 预期结果 / 适用端 / 用例等级 / 用例类型 / 备注) 与上传模板一致
2. 多步骤: 在 步骤描述 / 预期结果 单元格里用 换行 分隔, 一行一步, 不加 "【x】" 序号 (与上传解析对称)
3. 用例ID 列 (最后一列):
   - 空 = 新增用例
   - 填值 = 按 ID 更新 (ID 不存在会被拒绝)
4. 不删除任何用例; Excel 中没有的行 (DB 中有) = DB 保留原状
5. 所属分组: 用 | 分隔路径, 例 登录|账号|邮箱登录
6. 不要删除 / 重排 Sheet, 不要改 _meta (隐藏)"""


def _ensure_dv_sheet(
    wb: openpyxl.Workbook,
    level_labels: Optional[List[str]] = None,
    type_labels: Optional[List[str]] = None,
    platform_labels: Optional[List[str]] = None,
    all_group_paths: Optional[List[str]] = None,
) -> Dict[str, int]:
    """
    在 wb 里创建/重写隐藏的 _dv Sheet, 写入所有 DataValidation 下拉源数据.

    :returns: 各列行数 dict {"level": N, "type": N, "platform": N, "group": N}.
              0 表示该列没数据, 调用方据此决定是否挂对应 dv (0 就不挂).

    为啥要这个 sheet:
      老实现把下拉源写在主表的 Z/AA/AB/AC 隐藏列. 问题是用户在 Excel 里手填
      或从 B 列复制时, 容易把这些隐藏列也带上值, 解析端的 _is_empty_row
      会被"隐藏列残留值"干扰, 触发空行误判为非空.

      把下拉源集中到独立的隐藏 Sheet (_dv), 主表彻底干净 (最多 10 列业务列),
      不存在"用户能编辑的隐藏列", 根因消除. 跨 sheet 引用是 OOXML 标准,
      WPS / Excel / LibreOffice 都能解析.
    """
    if SHEET_DV in wb.sheetnames:
        dv_ws = wb[SHEET_DV]
        # 清空旧内容 (只清值, 不动列宽 / sheet_state)
        for row in dv_ws.iter_rows():
            for cell in row:
                cell.value = None
    else:
        dv_ws = wb.create_sheet(SHEET_DV)
    dv_ws.sheet_state = "hidden"

    # 写枚举值, 各列从 row 1 开始连续写, 跟 formula1 引用范围一致
    for col, labels in (
        (_DV_LEVEL_COL, level_labels or []),
        (_DV_TYPE_COL, type_labels or []),
        (_DV_PLATFORM_COL, platform_labels or []),
        (_DV_GROUP_COL, sorted(set(all_group_paths or []))),
    ):
        for i, v in enumerate(labels, start=1):
            dv_ws.cell(row=i, column=col, value=v)

    return {
        "level": len(level_labels or []),
        "type": len(type_labels or []),
        "platform": len(platform_labels or []),
        "group": len(set(all_group_paths or [])),
    }


def add_dropdowns_to_workbook(
    wb: openpyxl.Workbook,
    sheet_name: str,
    all_group_paths: Optional[List[str]] = None,
    level_labels: Optional[List[str]] = None,
    type_labels: Optional[List[str]] = None,
    platform_labels: Optional[List[str]] = None,
    max_row: int = 10000,
) -> None:
    """
    给已存在的 workbook / sheet 加 B/G/H/F 列下拉框 (DataValidation).
    复用逻辑: 导出 (build_workbook) 和下载模板 (download_case_template) 都调这个.

    参数:
      wb              openpyxl Workbook (已加载, 即将被 .save() / 序列化)
      sheet_name      主表 sheet 名 (导出叫 "用例数据", 模板叫 "template" / 默认)
      all_group_paths B 列下拉源 (全量目录路径, 形如 "根|登录|账号|邮箱登录")
                      None / 空 时跳过 B 列 dv
      level_labels    G 列下拉源 (用例等级 label, 来自 case_enum_config CASE_LEVEL)
      type_labels     H 列下拉源 (用例类型 label, 来自 case_enum_config CASE_TYPE)
      platform_labels F 列下拉源 (适用端 label, 来自 case_enum_config PLATFORM)
                      任一为 None / 空 时, 对应列不挂 dv (等同老 hardcode 缺失行为)
      max_row         dv 作用的最大行 (预留大值让用户继续追加新行也能用下拉)

    实现:
      1) 调用 _ensure_dv_sheet 在 _dv 隐藏 sheet 写所有下拉源 (主表 0 隐藏列)
      2) DataValidation formula1 走跨 sheet 引用: =_dv!$A$1:$A$N
      3) WPS / Excel / LibreOffice 都支持跨 sheet 引用的 list 类型 dv
    """
    if sheet_name not in wb.sheetnames:
        # 兜底: 用第一个 sheet
        sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]

    ends = _ensure_dv_sheet(
        wb,
        level_labels=level_labels,
        type_labels=type_labels,
        platform_labels=platform_labels,
        all_group_paths=all_group_paths,
    )
    level_end = ends["level"]
    type_end = ends["type"]
    platform_end = ends["platform"]
    group_end = ends["group"]

    # 1) G 列: 用例等级 -> _dv!A
    if level_end:
        dv_level = DataValidation(
            type="list",
            formula1=f"=_dv!$A$1:$A${level_end}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="非法值",
            error="请从下拉框中选择用例等级",
            promptTitle="用例等级",
            prompt=" / ".join(level_labels or []) or "请在配置中心维护用例等级",
        )
        ws.add_data_validation(dv_level)
        dv_level.add("G3:G{max_row}".format(max_row=max_row))

    # 2) H 列: 用例类型 -> _dv!B
    if type_end:
        dv_type = DataValidation(
            type="list",
            formula1=f"=_dv!$B$1:$B${type_end}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="非法值",
            error="请从下拉框中选择用例类型",
            promptTitle="用例类型",
            prompt=" / ".join(type_labels or []) or "请在配置中心维护用例类型",
        )
        ws.add_data_validation(dv_type)
        dv_type.add("H3:H{max_row}".format(max_row=max_row))

    # 3) F 列: 适用端 -> _dv!C
    if platform_end:
        dv_platform = DataValidation(
            type="list",
            formula1=f"=_dv!$C$1:$C${platform_end}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="非法值",
            error="请从下拉框中选择适用端",
            promptTitle="适用端",
            prompt=" / ".join(platform_labels or []) or "请在配置中心维护适用端",
        )
        ws.add_data_validation(dv_platform)
        dv_platform.add("F3:F{max_row}".format(max_row=max_row))

    # 4) B 列: 所属分组 (可选, 依赖调用方传 all_group_paths) -> _dv!D
    if group_end:
        dv_group = DataValidation(
            type="list",
            formula1=f"=_dv!$D$1:$D${group_end}",
            allow_blank=True,
            showDropDown=False,
            showErrorMessage=False,
            promptTitle="所属分组",
            prompt="从下拉选择目录路径 (多级用 | 分隔)",
        )
        ws.add_data_validation(dv_group)
        dv_group.add(f"B3:B{max_row}")




class ExportCaseService:
    """
    buf = ExportCaseService(
        scope_type, scope_id, case_dicts, group_path_map,
        label_map=None, all_group_paths=None,
        level_labels=None, type_labels=None, platform_labels=None,
    ).build_workbook()

    等级/类型/适用端的下拉项均由 controller 从 case_enum_config 加载后传入, 不写死.
    这样 admin 在配置中心增删枚举时, 导出的 Excel 下拉自动同步.
    """

    # 顺序固定, "用例ID" 放最后一列 (导出独有, 下载模板没有).
    DATA_COLUMNS: List[Tuple[str, str]] = [
        ("标题*", "case_name"),
        ("所属分组", "group_path"),
        ("前置条件", "case_setup"),
        ("步骤描述*", "action_cell"),
        ("预期结果*", "expected_result_cell"),
        ("适用端", "case_platform"),
        ("用例等级*", "case_level"),
        ("用例类型", "case_type"),
        ("备注", "case_mark"),
        ("用例ID", "case_id"),
    ]

    def __init__(
        self,
        scope_type: str,
        scope_id: int,
        case_dicts: List[Dict[str, Any]],
        group_path_map: Dict[int, str],
        label_map: Optional[Dict[str, str]] = None,
        all_group_paths: Optional[List[str]] = None,
        # 下拉项数据源: 业务枚举**不**写死, 由 controller 从 case_enum_config 加载后传入.
        # 顺序: case_config.sort asc. None / 空 时该列不挂 dv (等同老 hardcode 缺失行为).
        level_labels: Optional[List[str]] = None,
        type_labels: Optional[List[str]] = None,
        platform_labels: Optional[List[str]] = None,
    ):
        if scope_type not in ("library", "plan"):
            raise ValueError(f"scope_type 必须是 library 或 plan, 收到: {scope_type!r}")
        if len(case_dicts) > EXPORT_HARD_LIMIT:
            raise ValueError(
                f"导出量 {len(case_dicts)} 超过单次上限 {EXPORT_HARD_LIMIT}, "
                f"请先在页面用筛选缩小范围"
            )
        self.scope_type = scope_type
        self.scope_id = scope_id
        self.case_dicts = case_dicts
        self.group_path_map = group_path_map
        # case_config.value -> label. 缺 key 时回退到原 value (label_map.get(v, v)).
        self.label_map = label_map or {}
        # 全量目录路径列表 (平铺, 含根), 给 B 列 "所属分组" 做下拉选项.
        # None / 空 时不挂 B 列 dv, 用户继续手填.
        self.all_group_paths = sorted(set(all_group_paths or []))
        self.level_labels = list(level_labels or [])
        self.type_labels = list(type_labels or [])
        self.platform_labels = list(platform_labels or [])

    def build_workbook(self) -> BytesIO:
        # 以 file/用例模版.xlsx 为基底, 复用其样式/字体/排版, 改动量最小
        from file import TestCaseDemoFile
        wb = openpyxl.load_workbook(TestCaseDemoFile)
        ws = wb["template"]
        ws.title = SHEET_DATA

        # 替换 row 1: 标题 + 编辑引导 (覆盖原模板 A1 的 "导入模板" 标题)
        ws.cell(1, 1, _EXPORT_GUIDE)

        # 把模板 J2 原占位 "备注" 覆盖为 "用例ID". 导出独有的协议列, 解析端按此识别.
        ws.cell(2, len(self.DATA_COLUMNS), self.DATA_COLUMNS[-1][0])

        # 删 row 3 demo 行: 导出是真实数据, 留 demo 会被 PR-2 当成 INSERT 假数据.
        # 删完后真实数据从 row 3 开始写, 与 aioFileReader 默认 header_row+1 数据起点对齐.
        ws.delete_rows(3, 1)

        for row_idx, case in enumerate(self.case_dicts, start=3):
            self._write_data_row(ws, row_idx, case)

        # 给 "所属分组" (B 列) / "适用端" (F 列) / "用例等级" (G 列) / "用例类型" (H 列)
        # 加下拉框. 走统一的模块级 add_dropdowns_to_workbook: 它在 _dv 隐藏 Sheet
        # 写入下拉源, DataValidation 走跨 sheet 引用 (=_dv!$A$1:$A$N).
        # 主表彻底干净 (最多 10 列), 不再有 Z/AA/AB/AC 隐藏列.
        add_dropdowns_to_workbook(
            wb,
            sheet_name=SHEET_DATA,
            all_group_paths=self.all_group_paths,
            level_labels=self.level_labels,
            type_labels=self.type_labels,
            platform_labels=self.platform_labels,
            max_row=10000,
        )

        # 隐藏的 _meta Sheet: 跨 scope 防御 / 变更检测 / 版本号
        self._write_meta_sheet(wb)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _write_data_row(self, ws, row_idx: int, case: Dict[str, Any]) -> None:
        from utils.stepCellFormatter import format_steps_cell

        # scope 不同, 取的 id 字段不同; group_path_map 由 controller 按 scope 填好
        group_key = (
            case.get("plan_module_id") if self.scope_type == "plan" else case.get("module_id")
        )
        group_path = self.group_path_map.get(group_key, "") if group_key is not None else ""

        steps = case.get("case_sub_steps") or []
        action_cell = format_steps_cell(steps, "action")
        expected_result_cell = format_steps_cell(steps, "expected_result")

        case_level = case.get("case_level")
        case_type = case.get("case_type")
        case_platform = case.get("case_platform")

        values = [
            case.get("case_name"),
            group_path,
            case.get("case_setup"),
            action_cell,
            expected_result_cell,
            self.label_map.get(case_platform, case_platform),
            self.label_map.get(case_level, case_level),
            self.label_map.get(case_type, case_type),
            case.get("case_mark"),
            case.get("id"),
        ]
        for col_idx, v in enumerate(values, start=1):
            ws.cell(row=row_idx, column=col_idx, value=v)

    def _write_meta_sheet(self, wb) -> None:
        """
        隐藏的 _meta Sheet: scope / 导出快照 / 时间 / 版本号.

        - scope_type / scope_id: PR-3 commit 阶段做跨 scope 防御 (导出 plan:123, 不能上传到 plan:456)
        - case_ids_at_export: 解析端比对, 检测"导出后增删" 给前端做 warning
        - version: 协议版本号, 未来改格式时升级并 reject 老文件
        """
        case_ids = [c.get("id") for c in self.case_dicts if c.get("id") is not None]
        meta_ws = wb.create_sheet(SHEET_META)
        meta_ws["A1"], meta_ws["B1"] = "key", "value"
        rows = [
            ("scope_type", self.scope_type),
            ("scope_id", str(self.scope_id)),
            ("case_ids_at_export", ",".join(str(i) for i in case_ids)),
            ("exported_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("version", str(META_VERSION)),
        ]
        for i, (k, v) in enumerate(rows, start=2):
            meta_ws.cell(i, 1, k)
            meta_ws.cell(i, 2, v)
        meta_ws.column_dimensions["A"].width = 24
        meta_ws.column_dimensions["B"].width = 60
        meta_ws.sheet_state = "hidden"

    @property
    def case_count(self) -> int:
        return len(self.case_dicts)


def _walk_paths(
    id_to_node: Dict[int, Tuple[str, Optional[int]]],
    targets: List[int],
    max_depth: int = 50,
) -> Dict[int, str]:
    """
    通过 parent_id 链回溯构造 "A|B|C" 形式路径.

    用 "|" 而非 "/" 分隔, 跟 file/用例模版.xlsx 上传模板一致.
    mapper 层 (TestCaseMapper.build_module_path_map / PlanCaseMapper.build_plan_module_path_map)
    通过 from-import 复用本函数, 不在本服务内二次包装.
    """
    paths: Dict[int, str] = {}
    for nid in targets:
        parts: List[str] = []
        cur_id: Optional[int] = nid
        for _ in range(max_depth):
            node = id_to_node.get(cur_id) if cur_id is not None else None
            if node is None:
                break
            title, parent_id = node
            parts.append(title)
            if parent_id is None or parent_id == 0:
                break
            cur_id = parent_id
        if parts:
            paths[nid] = "|".join(reversed(parts))
    return paths
